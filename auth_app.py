"""
auth_app.py - Standalone Flask Auth Server (Port 5001)
Handles Signup with OTP email verification, Login, and Excel user storage.

SETUP:
  1. Create a .env file in this directory:
       MAIL_USER=yourgmail@gmail.com
       MAIL_PASSWORD=your_gmail_app_password   (NOT your normal password)
       SECRET_KEY=any_random_long_string
  2. To get a Gmail App Password:
       - Enable 2FA on your Google Account
       - Go to: Google Account → Security → 2-Step Verification → App passwords
       - Generate one for "Mail" and paste it here
  3. Run: python auth_app.py
  4. Open: http://localhost:5001
"""

import os
import smtplib
import ssl
import random
import string
import bcrypt
import openpyxl
import re
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, render_template, request, session, jsonify
from dotenv import load_dotenv
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production-secret-key-2026")

MAIL_USER     = os.environ.get("MAIL_USER", "")
MAIL_PASSWORD = os.environ.get("MAIL_PASSWORD", "")
SMTP_HOST     = "smtp.gmail.com"
SMTP_PORT     = 465

USERS_FILE    = Path(__file__).parent / "users.xlsx"
OTP_EXPIRY_MINUTES = 5
MAX_RESENDS   = 3

# In-memory OTP store: { email: { otp, expiry, resend_count } }
otp_store: dict = {}

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def _load_workbook():
    """Load or create the users Excel workbook."""
    if USERS_FILE.exists():
        return openpyxl.load_workbook(USERS_FILE)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["Email", "PasswordHash", "CreatedAt"])
    wb.save(USERS_FILE)
    return wb


def email_exists(email: str) -> bool:
    """Return True if email is already registered."""
    wb = _load_workbook()
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].lower() == email.lower():
            return True
    return False


def get_user_hash(email: str):
    """Return stored password hash for email, or None."""
    wb = _load_workbook()
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].lower() == email.lower():
            return row[1]
    return None


def save_user(email: str, password_plain: str):
    """Hash password and append new user row to Excel."""
    hashed = bcrypt.hashpw(password_plain.encode(), bcrypt.gensalt()).decode()
    wb = _load_workbook()
    ws = wb["Users"]
    ws.append([email.lower(), hashed, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(USERS_FILE)


# ---------------------------------------------------------------------------
# OTP helpers
# ---------------------------------------------------------------------------
def generate_otp() -> str:
    return "".join(random.choices(string.digits, k=6))


def store_otp(email: str, otp: str):
    otp_store[email.lower()] = {
        "otp": otp,
        "expiry": datetime.now() + timedelta(minutes=OTP_EXPIRY_MINUTES),
        "resend_count": otp_store.get(email.lower(), {}).get("resend_count", 0),
    }


def verify_otp(email: str, entered: str) -> tuple[bool, str]:
    """Returns (success, message)."""
    record = otp_store.get(email.lower())
    if not record:
        return False, "No OTP found. Please sign up again."
    if datetime.now() > record["expiry"]:
        otp_store.pop(email.lower(), None)
        return False, "OTP has expired. Please request a new one."
    if entered.strip() != record["otp"]:
        return False, "Invalid OTP. Please try again."
    # Valid — clean up
    otp_store.pop(email.lower(), None)
    return True, "OTP verified successfully!"


# ---------------------------------------------------------------------------
# OTP Email Template (Liquid Brokers - Premium Dark)
# ---------------------------------------------------------------------------
OTP_EMAIL_HTML = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        body { font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #050505; margin: 0; padding: 0; color: #ffffff; }
        .wrapper { background-color: #050505; padding: 40px 20px; text-align: center; }
        .container { max-width: 500px; margin: 0 auto; background-color: #111111; border-radius: 20px; border: 2px solid #d4af37; overflow: hidden; }
        .header { padding: 30px; background-color: #000000; }
        .brand { color: #d4af37; font-size: 22px; font-weight: 900; letter-spacing: 2px; text-transform: uppercase; }
        .content { padding: 40px; }
        .title { font-size: 14px; color: #aaaaaa; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 20px; }
        .otp-box { background-color: #d4af37; border-radius: 12px; padding: 25px; margin: 20px 0; }
        .otp-code { font-size: 40px; font-weight: 900; color: #000000; letter-spacing: 8px; }
        .warning { font-size: 12px; color: #888888; line-height: 1.5; margin-top: 20px; }
        .footer { padding: 20px; font-size: 10px; color: #555555; background-color: #000000; }
    </style>
</head>
<body>
    <div class="wrapper">
        <div class="container">
            <div class="header">
                <div class="brand">Virtual Life with AI</div>
            </div>
            <div class="content">
                <div class="title">Security Verification</div>
                <div class="otp-box">
                    <div class="otp-code">{{OTP}}</div>
                </div>
                <div class="warning">
                    This code is valid for <b>5 minutes</b>.<br>
                    Please do not share this with anyone.
                </div>
            </div>
            <div class="footer">
                &copy; 2026 Virtual Life with AI — Secure Access Portal
            </div>
        </div>
    </div>
</body>
</html>
"""

# ---------------------------------------------------------------------------
# Email sender
# ---------------------------------------------------------------------------
def send_otp_email(to_email: str, otp: str) -> tuple[bool, str]:
    """Send OTP email via Gmail SMTP formatted with Liquid Brokers Dark Theme."""
    if not MAIL_USER or not MAIL_PASSWORD:
        # Development fallback — print to console
        print(f"\n{'='*50}")
        print(f"  [DEV MODE] OTP for {to_email}: {otp}")
        print(f"  (Set MAIL_USER + MAIL_PASSWORD in .env to send real emails)")
        print(f"{'='*50}\n")
        return True, ""

    subject = f"{otp} is your verification code"
    
    # Create MIME message
    msg = MIMEMultipart("alternative")
    msg["From"]    = f"Virtual Life with AI <{MAIL_USER}>"
    msg["To"]      = to_email
    msg["Subject"] = subject

    # Plain text version
    text_body = f"Hello,\n\nYour Virtual Life verification code is: {otp}\n\nThis expires in 5 minutes.\n\n— Virtual Life with AI"
    
    # HTML version
    html_body = OTP_EMAIL_HTML.replace("{{OTP}}", otp)

    msg.attach(MIMEText(text_body, "plain"))
    msg.attach(MIMEText(html_body, "html"))

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context) as server:
            server.login(MAIL_USER, MAIL_PASSWORD)
            server.send_message(msg)
        return True, ""
    except smtplib.SMTPAuthenticationError:
        return False, "Email authentication failed. Check your MAIL_USER and MAIL_PASSWORD in .env."
    except Exception as e:
        return False, f"Failed to send email: {str(e)}"


# ---------------------------------------------------------------------------
# Input validation
# ---------------------------------------------------------------------------
def is_valid_email(email: str) -> bool:
    import re
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))


def is_valid_password(password: str) -> tuple[bool, str]:
    if len(password) < 8:
        return False, "Password must be at least 8 characters."
    return True, ""


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("auth.html")


@app.route("/status")
def status():
    """Return current login status for the homepage to consume."""
    return jsonify({
        "logged_in": "user_email" in session,
        "email": session.get("user_email", ""),
    })


@app.route("/signup", methods=["POST"])
def signup():
    data    = request.get_json()
    email   = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()
    confirm = (data.get("confirm_password") or "").strip()

    if not email or not password or not confirm:
        return jsonify(success=False, message="All fields are required.")
    if not is_valid_email(email):
        return jsonify(success=False, message="Please enter a valid email address.")
    if password != confirm:
        return jsonify(success=False, message="Passwords do not match.")
    ok, msg = is_valid_password(password)
    if not ok:
        return jsonify(success=False, message=msg)
    if email_exists(email):
        return jsonify(success=False, message="This email is already registered. Please log in.")

    # Store password temporarily in server-side session until OTP verified
    session["pending_email"]    = email.lower()
    session["pending_password"] = password  # will be hashed when OTP verified

    otp = generate_otp()
    store_otp(email, otp)
    success, err = send_otp_email(email, otp)
    if not success:
        return jsonify(success=False, message=err)

    return jsonify(success=True, message=f"OTP sent to {email}. Check your inbox (and spam folder).")


@app.route("/verify_otp", methods=["POST"])
def verify_otp_route():
    data  = request.get_json()
    email = (data.get("email") or session.get("pending_email") or "").strip()
    otp   = (data.get("otp") or "").strip()

    if not email or not otp:
        return jsonify(success=False, message="Email and OTP are required.")

    # Double-check email hasn't been registered between signup and verify steps
    if email_exists(email):
        session.pop("pending_email", None)
        session.pop("pending_password", None)
        return jsonify(success=False, message="This email is already registered.")

    ok, msg = verify_otp(email, otp)
    if not ok:
        return jsonify(success=False, message=msg)

    # OTP correct → save user
    password = session.pop("pending_password", None)
    session.pop("pending_email", None)
    if not password:
        return jsonify(success=False, message="Session expired. Please sign up again.")

    save_user(email, password)
    session["user_email"] = email.lower()
    return jsonify(success=True, message="Account created successfully! Welcome to Virtual Life with AI.")


@app.route("/resend-otp", methods=["POST"])
def resend_otp():
    data  = request.get_json()
    email = (data.get("email") or session.get("pending_email") or "").strip()

    if not email:
        return jsonify(success=False, message="No email found. Please start signup again.")

    record = otp_store.get(email.lower(), {})
    resend_count = record.get("resend_count", 0)
    if resend_count >= MAX_RESENDS:
        return jsonify(success=False, message="Maximum resend limit reached. Please start the signup process again.")

    otp = generate_otp()
    store_otp(email, otp)
    otp_store[email.lower()]["resend_count"] = resend_count + 1

    success, err = send_otp_email(email, otp)
    if not success:
        return jsonify(success=False, message=err)

    remaining = MAX_RESENDS - (resend_count + 1)
    return jsonify(success=True, message=f"New OTP sent to {email}. ({remaining} resend(s) remaining)")


@app.route("/login", methods=["POST"])
def login():
    data     = request.get_json()
    email    = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()

    if not email or not password:
        return jsonify(success=False, message="Email and password are required.")
    if not is_valid_email(email):
        return jsonify(success=False, message="Please enter a valid email address.")

    stored_hash = get_user_hash(email)
    if not stored_hash:
        return jsonify(success=False, message="No account found with this email. Please sign up.")

    if not bcrypt.checkpw(password.encode(), stored_hash.encode()):
        return jsonify(success=False, message="Incorrect password. Please try again.")

    session["user_email"] = email.lower()
    return jsonify(success=True, message=f"Login successful! Welcome back, {email}.")


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify(success=True, message="Logged out successfully.")


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("\n" + "="*55)
    print("  Virtual Life Auth Server running at http://localhost:5001")
    if not MAIL_USER:
        print("  ⚠  No .env found — OTPs will print to console (dev mode)")
    else:
        print(f"  ✉  SMTP configured for: {MAIL_USER}")
    print("="*55 + "\n")
    app.run(host="0.0.0.0", port=5001, debug=True)
