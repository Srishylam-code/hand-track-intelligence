"""
main.py â€” Hand Tracking + Auth System (single Flask server, port 5000)

  - Opening http://localhost:5000 shows the LOGIN page.
  - After a successful login/signup, the user sees the HandTrack dashboard.
  - Users are stored in users.xlsx with bcrypt-hashed passwords.
  - OTP is emailed via Gmail (or printed to console in dev mode).

Setup email (optional â€” skip for console-OTP dev mode):
  Copy .env.example â†’ .env and fill in MAIL_USER + MAIL_PASSWORD
"""

import argparse
import time
import sys
import cv2
import ssl
import smtplib
import random
import string
import threading
import webbrowser
import os
import re
import bcrypt
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, render_template, Response, request, jsonify, session, redirect, url_for

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from hand_tracker import HandTracker
from osc_sender import OSCSender
from hud_visualizer import HUDVisualizer

# â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
#  Flask App Config
# â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "handtrack-secret-key-2026-change-me")

MAIL_USER     = os.environ.get("MAIL_USER", "")
MAIL_PASSWORD = os.environ.get("MAIL_PASSWORD", "")
SMTP_HOST     = "smtp.gmail.com"
SMTP_PORT     = 587

USERS_FILE         = Path(__file__).parent / "users.xlsx"
OTP_EXPIRY_MINUTES = 5
MAX_RESENDS        = 3

# In-memory OTP store: { email: { otp, expiry, resend_count } }
otp_store: dict = {}

# Video streaming globals
output_frame = None
output_bytes = None
frame_lock   = threading.Lock()
camera_cmd   = None


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Excel helpers
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def _load_workbook():
    if USERS_FILE.exists():
        return openpyxl.load_workbook(USERS_FILE)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["Email", "PasswordHash", "CreatedAt"])
    wb.save(USERS_FILE)
    return wb


def email_exists(email: str) -> bool:
    wb = _load_workbook()
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].lower() == email.lower():
            return True
    return False


def get_user_hash(email: str):
    wb = _load_workbook()
    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].lower() == email.lower():
            return row[1]
    return None


def save_user(email: str, password_plain: str):
    hashed = bcrypt.hashpw(password_plain.encode(), bcrypt.gensalt()).decode()
    wb = _load_workbook()
    ws = wb["Users"]
    ws.append([email.lower(), hashed, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(USERS_FILE)


def seed_user_if_missing(email: str, password: str):
    """Pre-create a user account if it doesn't already exist."""
    if not email_exists(email):
        print(f"[AUTH] Pre-seeding account: {email}", flush=True)
        save_user(email, password)


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  OTP helpers
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def generate_otp() -> str:
    return "".join(random.choices(string.digits, k=6))


def store_otp(email: str, otp: str):
    otp_store[email.lower()] = {
        "otp":          otp,
        "expiry":       datetime.now() + timedelta(minutes=OTP_EXPIRY_MINUTES),
        "resend_count": otp_store.get(email.lower(), {}).get("resend_count", 0),
    }


def verify_otp(email: str, entered: str):
    record = otp_store.get(email.lower())
    if not record:
        return False, "No OTP found. Please sign up again."
    if datetime.now() > record["expiry"]:
        otp_store.pop(email.lower(), None)
        return False, "OTP has expired. Please request a new one."
    if entered.strip() != record["otp"]:
        return False, "Invalid OTP. Please try again."
    otp_store.pop(email.lower(), None)
    return True, "OTP verified successfully!"


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Email sender
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def send_otp_email(to_email: str, otp: str):
    if not MAIL_USER or not MAIL_PASSWORD:
        print(f"\n{'='*52}")
        print(f"  [DEV MODE] OTP for {to_email}: {otp}")
        print(f"  (Add MAIL_USER + MAIL_PASSWORD to .env for real emails)")
        print(f"{'='*52}\n", flush=True)
        return True, ""

    subject = "Your HandTrack Verification Code"
    body = (
        f"From: HandTrack Intelligence <{MAIL_USER}>\r\n"
        f"To: {to_email}\r\n"
        f"Subject: {subject}\r\n"
        f"Content-Type: text/plain; charset=utf-8\r\n\r\n"
        f"Hello,\r\n\r\n"
        f"Your one-time verification code is:\r\n\r\n"
        f"    {otp}\r\n\r\n"
        f"This code expires in {OTP_EXPIRY_MINUTES} minutes.\r\n"
        f"Do NOT share this code with anyone.\r\n\r\n"
        f"â€” HandTrack Intelligence\r\n"
    )
    try:
        ctx = ssl.create_default_context()
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.ehlo(); s.starttls(context=ctx); s.login(MAIL_USER, MAIL_PASSWORD)
            s.sendmail(MAIL_USER, to_email, body.encode("utf-8"))
        return True, ""
    except smtplib.SMTPAuthenticationError:
        return False, "Email auth failed. Check MAIL_USER / MAIL_PASSWORD in .env."
    except Exception as e:
        return False, f"Failed to send email: {e}"


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Validation helpers
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))


def is_valid_password(password: str):
    if len(password) < 8:
        return False, "Password must be at least 8 characters."
    return True, ""


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Auth guard decorator
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("auth_page"))
        return f(*args, **kwargs)
    return decorated


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Flask Routes â€” Pages
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/auth")
def auth_page():
    if "user_email" in session:
        return redirect(url_for("index"))
    return render_template("auth.html")


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Flask Routes â€” Auth API
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
@app.route("/status")
def status():
    return jsonify(logged_in="user_email" in session, email=session.get("user_email", ""))


@app.route("/login", methods=["POST"])
def login():
    data     = request.get_json()
    email    = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()

    if not email or not password:
        return jsonify(success=False, message="Email and password are required.")
    if not is_valid_email(email):
        return jsonify(success=False, message="Please enter a valid email address.")

    stored_hash = get_user_hash(email)
    if not stored_hash:
        return jsonify(success=False, message="No account found with this email. Please sign up.")
    if not bcrypt.checkpw(password.encode(), stored_hash.encode()):
        return jsonify(success=False, message="Incorrect password. Please try again.")

    session["user_email"] = email.lower()
    return jsonify(success=True, message=f"Login successful! Welcome back.")


@app.route("/signup", methods=["POST"])
def signup():
    data    = request.get_json()
    email   = (data.get("email") or "").strip()
    password = (data.get("password") or "").strip()
    confirm = (data.get("confirm_password") or "").strip()

    if not email or not password or not confirm:
        return jsonify(success=False, message="All fields are required.")
    if not is_valid_email(email):
        return jsonify(success=False, message="Please enter a valid email address.")
    if password != confirm:
        return jsonify(success=False, message="Passwords do not match.")
    ok, msg = is_valid_password(password)
    if not ok:
        return jsonify(success=False, message=msg)
    if email_exists(email):
        return jsonify(success=False, message="This email is already registered. Please log in.")

    session["pending_email"]    = email.lower()
    session["pending_password"] = password

    otp = generate_otp()
    store_otp(email, otp)
    ok, err = send_otp_email(email, otp)
    if not ok:
        return jsonify(success=False, message=err)

    return jsonify(success=True, message=f"OTP sent to {email}. Check your inbox.")


@app.route("/verify-otp", methods=["POST"])
def verify_otp_route():
    data  = request.get_json()
    email = (data.get("email") or session.get("pending_email") or "").strip()
    otp   = (data.get("otp") or "").strip()

    if not email or not otp:
        return jsonify(success=False, message="Email and OTP are required.")
    if email_exists(email):
        session.pop("pending_email", None); session.pop("pending_password", None)
        return jsonify(success=False, message="This email is already registered.")

    ok, msg = verify_otp(email, otp)
    if not ok:
        return jsonify(success=False, message=msg)

    password = session.pop("pending_password", None)
    session.pop("pending_email", None)
    if not password:
        return jsonify(success=False, message="Session expired. Please sign up again.")

    save_user(email, password)
    session["user_email"] = email.lower()
    return jsonify(success=True, message="Account created! Welcome to HandTrack.")


@app.route("/resend-otp", methods=["POST"])
def resend_otp():
    data  = request.get_json()
    email = (data.get("email") or session.get("pending_email") or "").strip()
    if not email:
        return jsonify(success=False, message="No email found. Please start signup again.")

    record       = otp_store.get(email.lower(), {})
    resend_count = record.get("resend_count", 0)
    if resend_count >= MAX_RESENDS:
        return jsonify(success=False, message="Maximum resend limit reached. Please start signup again.")

    otp = generate_otp()
    store_otp(email, otp)
    otp_store[email.lower()]["resend_count"] = resend_count + 1

    ok, err = send_otp_email(email, otp)
    if not ok:
        return jsonify(success=False, message=err)

    remaining = MAX_RESENDS - (resend_count + 1)
    return jsonify(success=True, message=f"New OTP sent. ({remaining} resend(s) remaining)")


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify(success=True, message="Logged out.")


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Flask Routes â€” Video / Camera
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def generate_frames():
    global output_bytes, frame_lock
    while True:
        with frame_lock:
            if output_bytes is None:
                encoded_image = None
            else:
                encoded_image = output_bytes
        
        if encoded_image is None:
            time.sleep(0.01)
            continue
        yield (b'--frame\r\nContent-Type: image/jpeg\r\n\r\n' + encoded_image + b'\r\n')


@app.route("/video_feed")
@login_required
def video_feed():
    return Response(generate_frames(), mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/set_camera", methods=["POST"])
@login_required
def set_camera():
    global camera_cmd
    data   = request.json
    source = data.get("source", "0")
    camera_cmd = int(source) if str(source).isdigit() else source
    return jsonify({"status": "success", "source": str(camera_cmd)})


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Flask Startup
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def start_flask():
    import logging
    log = logging.getLogger("werkzeug")
    log.setLevel(logging.ERROR)
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  CLI Args
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def parse_args():
    parser = argparse.ArgumentParser(description="Real-time hand tracking with OSC output")
    parser.add_argument("--ip",                   type=str,   default="127.0.0.1")
    parser.add_argument("--port",                 type=int,   default=9000)
    parser.add_argument("--camera",               type=int,   default=0)
    parser.add_argument("--max-hands",            type=int,   default=2, choices=[1, 2])
    parser.add_argument("--detection-confidence", type=float, default=0.7)
    parser.add_argument("--tracking-confidence",  type=float, default=0.5)
    parser.add_argument("--no-preview",           action="store_true")
    parser.add_argument("--width",                type=int,   default=1280)
    parser.add_argument("--height",               type=int,   default=720)
    return parser.parse_args()


# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
#  Main
# â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�â”�
def main():
    args = parse_args()

    # â”€â”€ Pre-seed user account â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    seed_user_if_missing("bandisrishylam77@gmail.com", "Srishylam@123")

    # â”€â”€ Initialize hand tracking components â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    tracker = HandTracker(
        max_num_hands=args.max_hands,
        min_detection_confidence=args.detection_confidence,
        min_tracking_confidence=args.tracking_confidence,
    )
    sender = OSCSender(ip=args.ip, port=args.port)
    hud    = HUDVisualizer()

    print(f"[INFO] Camera will start on standby (waiting for UI)...", flush=True)
    cap = None
    is_camera_running = False
    current_camera_src = args.camera

    print("=" * 55)
    print("  HandTrack â€” Login required at http://127.0.0.1:5000/auth")
    print("=" * 55)
    print(f"  Camera   : {args.camera} ({args.width}x{args.height})")
    print(f"  OSC      : {args.ip}:{args.port}")
    print(f"  Preview  : {'ON' if not args.no_preview else 'OFF'}")
    print("=" * 55)
    print("  Press 'q' to quit")
    print("=" * 55)

    flask_thread = threading.Thread(target=start_flask)
    flask_thread.daemon = True
    flask_thread.start()

    # Open browser directly to auth (login) page
    threading.Timer(1.5, lambda: webbrowser.open("http://127.0.0.1:5000/auth")).start()

    if not args.no_preview:
        cv2.namedWindow("Hand Tracking â†’ TouchDesigner", cv2.WINDOW_NORMAL)
        cv2.resizeWindow("Hand Tracking â†’ TouchDesigner", args.width, args.height)

    prev_time   = time.time()
    frame_count = 0
    elapsed     = 0.0
    global camera_cmd, output_bytes
    output_bytes = None

    try:
        while True:
            global output_frame, frame_lock
            if camera_cmd is not None:
                print(f"\n[INFO] Camera Command: {camera_cmd}", flush=True)
                if camera_cmd == "stop":
                    if cap is not None:
                        cap.release()
                    cap = None
                    is_camera_running = False
                    
                    # Clear output frame to clear the browser feed
                    with frame_lock:
                        output_frame = None
                elif camera_cmd == "start":
                    if cap is not None:
                        cap.release()
                    if sys.platform == "win32" and isinstance(current_camera_src, int):
                        cap = cv2.VideoCapture(current_camera_src, cv2.CAP_DSHOW)
                    else:
                        cap = cv2.VideoCapture(current_camera_src)
                    if cap and cap.isOpened():
                        cap.set(cv2.CAP_PROP_FRAME_WIDTH, args.width)
                        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, args.height)
                        is_camera_running = True
                else:
                    current_camera_src = camera_cmd
                    if is_camera_running:
                        if cap is not None:
                            cap.release()
                        if sys.platform == "win32" and isinstance(current_camera_src, int):
                            cap = cv2.VideoCapture(current_camera_src, cv2.CAP_DSHOW)
                        else:
                            cap = cv2.VideoCapture(current_camera_src)
                        if cap and cap.isOpened():
                            cap.set(cv2.CAP_PROP_FRAME_WIDTH, args.width)
                            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, args.height)
                camera_cmd = None

            if not is_camera_running or cap is None:
                time.sleep(0.1)
                continue

            ret, frame = cap.read()
            if not ret:
                time.sleep(0.01)
                continue

            frame = cv2.flip(frame, 1)
            hands = tracker.process(frame)
            sender.send_all(hands, max_hands=args.max_hands)

            current_time = time.time()
            elapsed      = current_time - prev_time

            # Draw HUD with scale-aware FPS
            fps_val = frame_count / elapsed if elapsed > 0 else 0
            frame = hud.draw(frame, hands, fps=fps_val)

            frame_count += 1

            if elapsed >= 1.0:
                fps         = frame_count / elapsed
                frame_count = 0
                prev_time   = current_time
                hand_info   = f"{len(hands)} hand(s)" if hands else "No hands"
                print(f"\r  FPS: {fps:.1f} | {hand_info}        ", end="", flush=True)

            with frame_lock:
                output_frame = frame.copy()
                success, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
                if success:
                    output_bytes = buffer.tobytes()

            if not args.no_preview:
                cv2.imshow("Hand Tracking â†’ TouchDesigner", frame)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                print("\n[INFO] Quit requested.")
                break

    except KeyboardInterrupt:
        print("\n[INFO] Interrupted by user.")
    finally:
        print("[INFO] Shutting down...")
        if cap is not None:
            cap.release()
        cv2.destroyAllWindows()
        tracker.close()
        sender.close()
        print("[INFO] Done.")


if __name__ == "__main__":
    main()
